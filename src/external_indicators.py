"""
External SSN-published indicators (from the official Excel report).

Some SSN-published indicators are not derivable from the balance parquet:

  - B (Cantidad de Juicios)            – a count, not a balance line
  - G (% Superávit / Capital Requerido) – regulatory calc (Res. 38.708)
  - H (Disp+Inv / Compromisos Exigibles) – aggregate definition we don't have

Plus J (Siniestralidad) requires reserve-change adjustments whose exact SSN
treatment we don't have authoritative documentation for; we serve the
official Excel value rather than approximate.

The Excel reports live in Input/ssn_<YYYYMM>_indicadores_mercado*.xlsx and
each corresponds to one SSN quarter. This module discovers all such files,
parses them, and exposes per-quarter indicator tables keyed by `cod_cia`
via name-matching against the parquet.
"""
from __future__ import annotations
import re
import unicodedata
from functools import lru_cache
from pathlib import Path

import openpyxl
import pandas as pd


INPUT_DIR = Path(__file__).parent.parent / "Input"

# YYYYMM month → SSN quarter suffix
_MONTH_TO_QUARTER = {
    "03": "Q1",
    "06": "Q2",
    "09": "Q3",
    "12": "Q4",
}

# Indicator code → column index in each Excel sheet
_SHEET1_COLS = {3: "A", 4: "B", 5: "C", 6: "D", 7: "D'", 8: "E", 9: "F", 10: "G", 11: "H"}
_SHEET2_COLS = {3: "I", 4: "J", 5: "K", 6: "L", 7: "M", 8: "N"}

# Indicators we serve from Excel rather than computing from parquet
EXTERNAL_CODES: list[str] = ["B", "G", "H", "J"]

# Static metadata for external indicators (name, category)
EXTERNAL_DEFS = [
    ("B", "Cantidad de Juicios",                     "patrimonial"),
    ("G", "% Superávit / Capital Requerido",          "patrimonial"),
    ("H", "(Disp+Inv) / Compromisos Exigibles",       "patrimonial"),
    ("J", "% Siniestros Netos Devengados / Primas Netas Devengadas", "gestion"),
]


# ── name normalization & matching ─────────────────────────────────────────────

_ABBREV = {"FED": "FEDERACION", "CIA": "COMPANIA", "SEG": "SEGUROS"}
_NOISE = {
    "S", "A", "U", "SAU", "SA", "SOCIEDAD", "ANONIMA", "UNIPERSONAL",
    "COMPANIA", "COMPANY", "DE", "DEL", "LA", "EL", "LIMITADA", "LTDA",
    "LTD", "SEGUROS", "PATRIMONIALES", "PERSONAS",
}


def _tokens(name: str | None) -> set[str]:
    if not name:
        return set()
    s = str(name).strip()
    s = "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
    s = s.upper()
    s = re.sub(r"[.,]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    parts = [_ABBREV.get(p, p) for p in s.split()]
    return {p for p in parts if len(p) > 1 and p not in _NOISE}


def _match_company(excel_tokens: set[str], parquet_df: pd.DataFrame) -> str | None:
    """Return cod_cia of best parquet match, or None below confidence threshold."""
    if not excel_tokens:
        return None
    best, best_score = None, 0.0
    for _, p in parquet_df.iterrows():
        pt = p["__tokens"]
        if not pt:
            continue
        inter = excel_tokens & pt
        if not inter:
            continue
        score = 100 * len(inter) / len(excel_tokens) + 10 * len(inter) / len(pt)
        if score > best_score:
            best_score = score
            best = p["cod_cia"]
    return best if best_score >= 50 else None


# ── Excel parser ──────────────────────────────────────────────────────────────

def _list_excel_files() -> dict[str, Path]:
    """Map SSN quarter (e.g. '2025-Q4') → Excel path."""
    out: dict[str, Path] = {}
    pat = re.compile(r"ssn_(\d{4})(\d{2})_indicadores_mercado.*\.xlsx$", re.I)
    for f in INPUT_DIR.glob("ssn_*.xlsx"):
        m = pat.match(f.name)
        if not m:
            continue
        q = _MONTH_TO_QUARTER.get(m.group(2))
        if q:
            out[f"{m.group(1)}-{q}"] = f
    return out


def _parse_sheet(ws, col_map: dict[int, str]) -> list[dict]:
    rows = []
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if not isinstance(r[0], (int, float)) or not r[2]:
            continue  # skip headers and subtotal rows
        row: dict = {"denom": str(r[2]).strip()}
        for idx, code in col_map.items():
            v = r[idx] if idx < len(r) else None
            row[code] = float(v) if isinstance(v, (int, float)) else None
        rows.append(row)
    return rows


@lru_cache(maxsize=16)
def _parse_excel_cached(path_str: str, mtime: float) -> pd.DataFrame:
    """Inner cache — mtime arg busts the cache when the file is updated."""
    wb = openpyxl.load_workbook(path_str, data_only=True)
    s1 = pd.DataFrame(_parse_sheet(wb["1 Indicadores Patrimoniales"], _SHEET1_COLS))
    s2 = pd.DataFrame(_parse_sheet(wb["2 Indicadores Gestion"], _SHEET2_COLS))
    return pd.merge(s1, s2, on="denom", how="outer")


def parse_excel(quarter: str) -> pd.DataFrame:
    """
    Wide DataFrame with columns: denom, A, B, C, D, D', E, F, G, H, I, J, K, L, M, N.
    Empty DataFrame if no Excel exists for the quarter.
    """
    files = _list_excel_files()
    if quarter not in files:
        return pd.DataFrame()
    p = files[quarter]
    return _parse_excel_cached(str(p), p.stat().st_mtime)


# ── public API ────────────────────────────────────────────────────────────────

def available_quarters() -> list[str]:
    """SSN quarters for which we have an Excel file."""
    return sorted(_list_excel_files().keys())


def load_external_long(parquet_df: pd.DataFrame, quarter: str,
                        codes: list[str] | None = None) -> pd.DataFrame:
    """
    Long-format external indicators for `quarter`, matched to parquet cod_cia.

    Returns columns: cod_cia, indicator, value. Returns empty DataFrame if no
    Excel covers the requested quarter.
    """
    excel = parse_excel(quarter)
    if excel.empty:
        return pd.DataFrame(columns=["cod_cia", "indicator", "value"])

    if codes is None:
        codes = EXTERNAL_CODES

    # Build name → cod_cia lookup from parquet (scoped to the quarter)
    pq = (
        parquet_df.loc[parquet_df["quarter"] == quarter, ["cod_cia", "razon_social"]]
                  .drop_duplicates("cod_cia").copy()
    )
    pq["__tokens"] = pq["razon_social"].apply(_tokens)

    rows = []
    for _, e in excel.iterrows():
        cod = _match_company(_tokens(e["denom"]), pq)
        if cod is None:
            continue
        for code in codes:
            if code not in e:
                continue
            v = e[code]
            if v is None or pd.isna(v):
                continue
            rows.append({"cod_cia": cod, "indicator": code, "value": float(v)})
    return pd.DataFrame(rows)


def match_report(parquet_df: pd.DataFrame, quarter: str) -> dict:
    """
    Debug helper: report on the name-matching quality for `quarter`.
    Returns counts and the list of unmatched Excel rows.
    """
    excel = parse_excel(quarter)
    if excel.empty:
        return {"available": False}

    pq = (
        parquet_df.loc[parquet_df["quarter"] == quarter, ["cod_cia", "razon_social"]]
                  .drop_duplicates("cod_cia").copy()
    )
    pq["__tokens"] = pq["razon_social"].apply(_tokens)

    matched, unmatched = 0, []
    for _, e in excel.iterrows():
        cod = _match_company(_tokens(e["denom"]), pq)
        if cod:
            matched += 1
        else:
            unmatched.append(e["denom"])
    return {
        "available": True,
        "matched": matched,
        "total_excel": len(excel),
        "unmatched": unmatched,
    }
